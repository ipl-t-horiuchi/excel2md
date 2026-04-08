import json
import time
import boto3
from botocore.config import Config
from botocore.exceptions import ClientError
from openpyxl.utils import get_column_letter
import openpyxl
import base64
import os
import tempfile
from urllib.parse import unquote_plus
import uuid as uuid_module


# ─────────────────────────────────────────────
# メインハンドラ
# ─────────────────────────────────────────────

def _http_method(event: dict):
    if event.get("httpMethod"):
        return event["httpMethod"]
    return (event.get("requestContext", {}).get("http") or {}).get("method")


def _parse_json_body(event: dict) -> dict:
    """API Gateway は body を Base64 エンコードすることがある（isBase64Encoded）。"""
    raw = event.get("body")
    if raw is None:
        return {}
    if event.get("isBase64Encoded"):
        raw = base64.b64decode(raw).decode("utf-8")
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8")
    raw = raw.strip()
    if not raw:
        return {}
    return json.loads(raw)


def lambda_handler(event, context):
    # 非同期 reconvert 実行（自己呼び出しで届く）
    if event.get("_reconvert"):
        return _do_reconvert(event)

    method = _http_method(event)
    if method is not None:
        return _route_api(event, method)
    if "Records" in event:
        return _handle_s3_event(event)
    return {"statusCode": 400, "body": json.dumps({"error": "不明なイベント種別です"})}


def _route_api(event, method):
    if method == "OPTIONS":
        return _cors(200, {})

    path = event.get("path") or event.get("rawPath") or "/"

    if "/presign" in path and method == "POST":
        return _handle_presign()
    # 「/reconvert」は /reconvert-cancel にもマッチするため、先に長いパスを判定する
    if "/reconvert-cancel" in path and method == "POST":
        return _handle_reconvert_cancel(event)
    if "/reconvert-status" in path and method == "GET":
        return _handle_reconvert_status(event)
    if "/reconvert" in path and method == "POST":
        return _handle_reconvert_api(event)

    return _cors(404, {"error": "見つかりません"})


# ─────────────────────────────────────────────
# API: POST /presign
#   AI 再変換用にアップロード URL を発行する
# ─────────────────────────────────────────────

def _handle_presign():
    s3 = boto3.client("s3")
    input_bucket = os.environ["INPUT_BUCKET"]
    job_id = str(uuid_module.uuid4())
    input_key = f"jobs/{job_id}.xlsx"

    upload_url = s3.generate_presigned_url(
        "put_object",
        Params={
            "Bucket": input_bucket,
            "Key": input_key,
            "ContentType": "application/octet-stream",
        },
        ExpiresIn=300,
    )
    return _cors(200, {"uploadUrl": upload_url, "jobId": job_id})


# ─────────────────────────────────────────────
# API: POST /reconvert
#   { jobId, sheetNames[] } → { reconvertId }
#   Lambda を非同期に自己呼び出しして Bedrock 変換を実行
# ─────────────────────────────────────────────

def _handle_reconvert_api(event):
    try:
        body = _parse_json_body(event)
    except json.JSONDecodeError as e:
        print(f"reconvert JSON parse error: {e} body_preview={repr((event.get('body') or '')[:200])}")
        return _cors(400, {"error": "JSON 本文が不正です"})
    job_id = body.get("jobId", "").strip()
    sheet_names = body.get("sheetNames") or []
    if not job_id or not sheet_names:
        return _cors(400, {"error": "jobId と sheetNames が必要です"})

    reconvert_id = str(uuid_module.uuid4())
    output_bucket = os.environ["OUTPUT_BUCKET"]
    s3 = boto3.client("s3")

    # 処理中マーカー
    s3.put_object(
        Bucket=output_bucket,
        Key=f"jobs/{reconvert_id}.reconvert.processing",
        Body=b"",
        ContentType="application/octet-stream",
    )

    lam = boto3.client("lambda")
    lam.invoke(
        FunctionName=os.environ["AWS_LAMBDA_FUNCTION_NAME"],
        InvocationType="Event",
        Payload=json.dumps({
            "_reconvert": True,
            "jobId": job_id,
            "reconvertId": reconvert_id,
            "sheetNames": sheet_names,
        }).encode(),
    )
    print(f"Reconvert requested: jobId={job_id} reconvertId={reconvert_id} sheets={sheet_names}")
    return _cors(200, {"reconvertId": reconvert_id})


# ─────────────────────────────────────────────
# API: POST /reconvert-cancel
#   { reconvertId } → S3 にキャンセルマーカーを置き、ワーカーが検知して停止する
# ─────────────────────────────────────────────

def _reconvert_cancel_key(reconvert_id: str) -> str:
    return f"jobs/{reconvert_id}.reconvert.cancel"


def _handle_reconvert_cancel(event):
    try:
        body = _parse_json_body(event)
    except json.JSONDecodeError:
        return _cors(400, {"error": "JSON 本文が不正です"})
    reconvert_id = (body.get("reconvertId") or "").strip()
    if not reconvert_id:
        return _cors(400, {"error": "reconvertId が必要です"})
    output_bucket = os.environ["OUTPUT_BUCKET"]
    s3 = boto3.client("s3")
    s3.put_object(
        Bucket=output_bucket,
        Key=_reconvert_cancel_key(reconvert_id),
        Body=b"",
        ContentType="application/octet-stream",
    )
    print(f"Reconvert cancel requested: reconvertId={reconvert_id}")
    return _cors(200, {"ok": True})


def _is_reconvert_cancel_requested(s3, output_bucket: str, reconvert_id: str) -> bool:
    try:
        s3.head_object(Bucket=output_bucket, Key=_reconvert_cancel_key(reconvert_id))
        return True
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code", "")
        if code in ("404", "NoSuchKey", "NotFound"):
            return False
        raise


def _cleanup_reconvert_cancelled(s3, input_bucket: str, output_bucket: str, job_id: str, reconvert_id: str):
    """キャンセル確定時: 処理中マーカー削除・入力 xlsx 削除（キャンセルマーカーは残す）"""
    try:
        s3.delete_object(Bucket=output_bucket, Key=f"jobs/{reconvert_id}.reconvert.processing")
    except Exception:
        pass
    try:
        s3.delete_object(Bucket=input_bucket, Key=f"jobs/{job_id}.xlsx")
    except Exception:
        pass


# ─────────────────────────────────────────────
# API: GET /reconvert-status?reconvertId=xxx
# ─────────────────────────────────────────────

def _handle_reconvert_status(event):
    s3 = boto3.client("s3")
    params = event.get("queryStringParameters") or {}
    reconvert_id = params.get("reconvertId", "").strip()
    if not reconvert_id:
        return _cors(400, {"error": "reconvertId が必要です"})

    output_bucket = os.environ["OUTPUT_BUCKET"]
    json_key = f"jobs/{reconvert_id}.reconvert.json"

    try:
        obj = s3.get_object(Bucket=output_bucket, Key=f"jobs/{reconvert_id}.reconvert.error")
        msg = obj["Body"].read().decode("utf-8")
        return _cors(200, {"status": "error", "error": msg})
    except Exception:
        pass

    try:
        s3.head_object(Bucket=output_bucket, Key=json_key)
        download_url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": output_bucket, "Key": json_key},
            ExpiresIn=3600,
        )
        return _cors(200, {"status": "done", "downloadUrl": download_url})
    except Exception:
        pass

    try:
        s3.head_object(Bucket=output_bucket, Key=_reconvert_cancel_key(reconvert_id))
        return _cors(200, {"status": "cancelled"})
    except Exception:
        pass

    return _cors(200, {"status": "processing"})


# ─────────────────────────────────────────────
# 非同期 reconvert 実行（Lambda 自己呼び出し）
#   INPUT_BUCKET から xlsx をダウンロードし Bedrock で変換
#   結果を { sheets: [{name, markdown}] } JSON として OUTPUT_BUCKET に保存
# ─────────────────────────────────────────────

def _do_reconvert(event):
    job_id = event["jobId"]
    reconvert_id = event["reconvertId"]
    sheet_names: list[str] = event["sheetNames"]
    t0 = time.perf_counter()
    print(f"Reconvert start: jobId={job_id} reconvertId={reconvert_id} sheets={sheet_names}")

    s3 = boto3.client("s3")
    input_bucket = os.environ["INPUT_BUCKET"]
    output_bucket = os.environ["OUTPUT_BUCKET"]
    xlsx_key = f"jobs/{job_id}.xlsx"
    json_key = f"jobs/{reconvert_id}.reconvert.json"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        excel_path = tmp.name

    try:
        if _is_reconvert_cancel_requested(s3, output_bucket, reconvert_id):
            _cleanup_reconvert_cancelled(s3, input_bucket, output_bucket, job_id, reconvert_id)
            print(f"[reconvert] cancelled before download reconvertId={reconvert_id}")
            return {"statusCode": 200}

        try:
            s3.download_file(input_bucket, xlsx_key, excel_path)
        except ClientError as e:
            code = e.response.get("Error", {}).get("Code", "")
            if code in ("404", "NoSuchKey", "NotFound"):
                raise RuntimeError(
                    f"入力バケットに xlsx が見つかりません: s3://{input_bucket}/{xlsx_key}"
                ) from e
            raise

        if _is_reconvert_cancel_requested(s3, output_bucket, reconvert_id):
            _cleanup_reconvert_cancelled(s3, input_bucket, output_bucket, job_id, reconvert_id)
            print(f"[reconvert] cancelled after download reconvertId={reconvert_id}")
            return {"statusCode": 200}

        print(f"[reconvert] xlsx downloaded in {time.perf_counter() - t0:.2f}s")
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        to_convert = [
            (name, _extract_sheet_data(wb[name]))
            for name in sheet_names
            if name in wb.sheetnames
        ]

        # キャンセル検知のためシートは順次処理（並列だと未投入分を止めにくい）
        results: dict[str, str] = {}
        for name, data in to_convert:
            if _is_reconvert_cancel_requested(s3, output_bucket, reconvert_id):
                print(f"[reconvert] cancelled before sheet={name!r} reconvertId={reconvert_id}")
                _cleanup_reconvert_cancelled(s3, input_bucket, output_bucket, job_id, reconvert_id)
                return {"statusCode": 200}
            try:
                results[name] = _call_bedrock_claude(data, name)
                print(f"[reconvert] Bedrock done: sheet='{name}' in {time.perf_counter() - t0:.2f}s")
            except Exception as e:
                results[name] = f"（AI変換エラー: {e}）"
                print(f"[reconvert] Bedrock error: sheet='{name}' {e}")

        if _is_reconvert_cancel_requested(s3, output_bucket, reconvert_id):
            _cleanup_reconvert_cancelled(s3, input_bucket, output_bucket, job_id, reconvert_id)
            print(f"[reconvert] cancelled after sheets reconvertId={reconvert_id}")
            return {"statusCode": 200}

        # 結果を JSON で保存
        sheets_json = [
            {"name": name, "markdown": results[name]}
            for name in sheet_names
            if name in results
        ]
        s3.put_object(
            Bucket=output_bucket,
            Key=json_key,
            Body=json.dumps({"sheets": sheets_json}, ensure_ascii=False).encode("utf-8"),
            ContentType="application/json; charset=utf-8",
        )

        # 処理中マーカー削除・xlsx クリーンアップ
        s3.delete_object(Bucket=output_bucket, Key=f"jobs/{reconvert_id}.reconvert.processing")
        try:
            s3.delete_object(Bucket=input_bucket, Key=xlsx_key)
        except Exception:
            pass

        print(f"Reconvert done: jobId={job_id} in {time.perf_counter() - t0:.2f}s")
        return {"statusCode": 200}

    except Exception as e:
        print(f"Reconvert error: {e}")
        s3.put_object(
            Bucket=output_bucket,
            Key=f"jobs/{reconvert_id}.reconvert.error",
            Body=str(e).encode("utf-8"),
            ContentType="text/plain",
        )
        s3.delete_object(Bucket=output_bucket, Key=f"jobs/{reconvert_id}.reconvert.processing")
        raise
    finally:
        os.unlink(excel_path)


# ─────────────────────────────────────────────
# S3 イベントハンドラ（ノーオプ）
#   初回変換はフロントエンドで実行するため、S3 トリガーは何もしない
# ─────────────────────────────────────────────

def _handle_s3_event(event):
    record = event["Records"][0]["s3"]
    bucket = record["bucket"]["name"]
    key = unquote_plus(record["object"]["key"])
    print(f"S3 event received (no-op): s3://{bucket}/{key}")
    return {"statusCode": 200, "body": "no-op"}


# ─────────────────────────────────────────────
# Excel シートデータ抽出（reconvert 用）
# ─────────────────────────────────────────────

def _extract_sheet_data(ws) -> dict:
    merged_cells = [str(m) for m in ws.merged_cells.ranges]
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                cells.append({
                    "pos": cell.coordinate,
                    "row": cell.row,
                    "col": get_column_letter(cell.column),
                    "value": str(cell.value),
                })

    images = []
    for img in getattr(ws, "_images", []):
        try:
            img_bytes = img._data()
            images.append({
                "anchor": str(img.anchor),
                "data": base64.b64encode(img_bytes).decode("utf-8"),
                "media_type": "image/png" if img_bytes[:4] == b"\x89PNG" else "image/jpeg",
            })
        except Exception as e:
            print(f"Image extraction error: {e}")

    return {"merged_cells": merged_cells, "cells": cells, "images": images}


# ─────────────────────────────────────────────
# Bedrock Claude 呼び出し（reconvert 用）
# ─────────────────────────────────────────────

def _call_bedrock_claude(sheet_data: dict, sheet_name: str) -> str:
    bedrock = boto3.client(
        "bedrock-runtime",
        region_name=os.environ.get("BEDROCK_REGION", "ap-northeast-1"),
        config=Config(read_timeout=600, connect_timeout=10),
    )

    cell_text = "\n".join(f"  {c['pos']}: {c['value']}" for c in sheet_data["cells"])
    merged_text = (
        "\n".join(f"  - {m}" for m in sheet_data["merged_cells"])
        if sheet_data["merged_cells"]
        else "  （結合なし）"
    )

    prompt = f"""以下はExcelシート「{sheet_name}」から抽出したデータです。

## 結合セル範囲
{merged_text}

## セルデータ（座標: 値）
{cell_text}

## 変換ルール（必ず守ること）
- 内容を一切要約・省略しないこと。すべての情報を出力すること
- 元の文章・文言を変更しないこと（誤字脱字も含めて原文のまま）
- 表構造は結合セル情報を考慮してMarkdownテーブルに変換すること
- 項番や階層構造はMarkdownの見出し・リストで表現すること
- エクセル方眼紙による細かいセル分割や空白セルは無視し、内容を正しくまとめること
- 表の中に表がある場合は、ネスト構造またはセクション分けで表現すること
- 画像がある場合は「![図: 内容の説明](添付画像)」として記載すること

上記のExcelシートの内容をMarkdown形式に変換してください。"""

    content: list = [{"type": "text", "text": prompt}]
    for img in sheet_data["images"]:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": img["media_type"],
                "data": img["data"],
            },
        })
    if sheet_data["images"]:
        content.append({
            "type": "text",
            "text": "シート内に埋め込まれた画像も確認し、内容や図の説明をMarkdownに含めてください。",
        })

    body = json.dumps({
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": 8192,
        "messages": [{"role": "user", "content": content}],
    })

    model_id = os.environ.get(
        "BEDROCK_MODEL_ID", "jp.anthropic.claude-sonnet-4-5-20250929-v1:0"
    )
    print(
        f"Bedrock invoke start: sheet={sheet_name!r} modelId={model_id} "
        f"body_bytes={len(body.encode('utf-8'))}"
    )
    t_br = time.perf_counter()
    response = bedrock.invoke_model_with_response_stream(modelId=model_id, body=body)
    chunks: list[str] = []
    for event in response["body"]:
        chunk = event.get("chunk")
        if chunk:
            data = json.loads(chunk["bytes"].decode("utf-8"))
            if data.get("type") == "content_block_delta":
                delta = data.get("delta", {})
                if delta.get("type") == "text_delta":
                    chunks.append(delta.get("text", ""))
    print(f"Bedrock invoke done: sheet={sheet_name!r} in {time.perf_counter() - t_br:.2f}s")
    return "".join(chunks)


# ─────────────────────────────────────────────
# CORS ヘッダー付きレスポンス
# ─────────────────────────────────────────────

def _cors(status_code: int, body: dict) -> dict:
    return {
        "statusCode": status_code,
        "headers": {
            "Content-Type": "application/json; charset=utf-8",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": (
                "Content-Type,Authorization,X-Amz-Date,X-Api-Key,X-Amz-Security-Token"
            ),
            "Access-Control-Allow-Methods": "GET,POST,OPTIONS",
            "Access-Control-Max-Age": "86400",
        },
        "body": json.dumps(body, ensure_ascii=False),
    }
